from process import *
from matplotlib import pyplot as plt
import os
from openpyxl import Workbook, load_workbook
import glob

"""
label_path = 'C:/Users/User/Documents/материалы ВИШ/АСЭ/chip_analysis/data/Xrays/5/14270175-x1(12).jpg'
if os.path.isfile(label_path):
    print(f"Изображение существует")
else:
    print(f"НЕ существует.")



etalon_img = cv2.imdecode(np.fromfile(label_path, dtype=np.uint8), cv2.IMREAD_GRAYSCALE)
etalon_cut_data = get_window(etalon_img)
etalon_cut_data = etalon_cut_data['img_cutted']
# etalon_cut_data = get_window(etalon_img)['img_cutted']

# Выводим изображение эталона с помощью plt
fig = plt.figure(figsize=(5, 4))
fig.add_subplot(1, 1, 1)
plt.imshow(etalon_cut_data, 'gray')
plt.title('Эталонное изображение')
plt.show()
"""

def update_excel(file_path, folder_name, test_image_name, image_difference):
    """
    Функция для создания или обновления Excel-файла с указанными данными.
    """
    try:
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Партия", "Снимок", "Статус"])  # заголовки


        # Проверяем, есть ли уже такие данные в файле
        found = False
        for row in ws.iter_rows(min_row=2, max_col=3, max_row=ws.max_row):
            if row[0].value == folder_name and row[1].value == test_image_name:
                row[2].value = image_difference
                found = True
                break

        if not found:
            ws.append([folder_name, test_image_name, image_difference])

        wb.save(file_path)
    except PermissionError:
        print("Ошибка: Файл Excel открыт. Пожалуйста, закройте файл и повторите попытку.")

# Пример использования
file_path = r"C:\Users\User\Documents\материалы ВИШ\АСЭ\chip_analysis\results2.xlsx"
folder_name = "3"
test_image_name = "TestImage1.jpg"
image_difference = "Различия есть"
# image_difference = "Различий нет"

update_excel(file_path, folder_name, test_image_name, image_difference)