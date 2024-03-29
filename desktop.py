import sys
import os
from openpyxl import Workbook, load_workbook
from PyQt5.QtGui import QPixmap, QImage, QFont, qRgb
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QGroupBox, QPushButton, QVBoxLayout, QFileDialog, \
    QScrollArea, QHBoxLayout, QGraphicsPixmapItem, QGraphicsScene, QGraphicsView, QCheckBox
from PyQt5.QtCore import Qt
from PyQt5.QtCore import QUrl
from PyQt5.QtGui import QDesktopServices

from process import *

# Была проблема в комфликте версий. Решение - указать на использование плагинов Qt, используемых PyQt5.
from pathlib import Path
import PyQt5
os.environ["QT_QPA_PLATFORM_PLUGIN_PATH"] = os.fspath(
    Path(PyQt5.__file__).resolve().parent / "Qt5" / "plugins")


class ImageComparisonApp(QWidget):
    def __init__(self):
        super().__init__()

        self.folder_path = ""
        self.label_path = ""
        self.test_image_paths = []
        self.images_loaded = False  # Флаг для отслеживания состояния загрузки изображений
        self.excel_file = r"C:\Users\User\Documents\материалы ВИШ\АСЭ\chip_analysis\results.xlsx"

        self.initUI()

    def initUI(self):
        layout = QVBoxLayout()
        print("FOLD | LABEL | TEST_IMAGE | STATUS")

        # Увеличим размер шрифта
        font = QFont()
        font.setPointSize(12)

        h_layout = QHBoxLayout()


        self.folder_label = QLabel("Выберите папку с изображениями:")
        self.folder_label.setFont(font)
        h_layout.addWidget(self.folder_label)


        self.results_button = QPushButton("Результаты сеансов")
        self.results_button.setFont(font)
        self.results_button.setStyleSheet(
            "QPushButton {"
            "background-color: #66c266;"  # Зеленый цвет кнопки
            "border: 1px solid #339933;"
            "color: white;"
            "border-radius: 5px;"
            "}"
            "QPushButton:hover {"
            "background-color: #4caf50;"  # Цвет кнопки при наведении курсора
            "}"
            "QPushButton:pressed {"
            "background-color: #388e3c;"  # Цвет кнопки при нажатии
            "}"
        )
        button_size = self.results_button.fontMetrics().boundingRect(self.results_button.text()).size() # Установка размера кнопки по размеру текста внутри
        self.results_button.setFixedSize(button_size.width()+10, button_size.height()+5)
        self.results_button.clicked.connect(self.openExcelFile)
        h_layout.addWidget(self.results_button)

        # Добавление горизонтального контейнера в вертикальный контейнер layout
        layout.addLayout(h_layout)


        self.folder_button = QPushButton("Выбрать папку")
        self.folder_button.setFont(font)
        self.folder_button.clicked.connect(self.selectFolder)
        layout.addWidget(self.folder_button)


        self.reference_label = QLabel("Выберите эталон:")
        self.reference_label.setFont(font)
        layout.addWidget(self.reference_label)

        self.reference_button = QPushButton("Выбрать эталон")
        self.reference_button.setFont(font)
        self.reference_button.clicked.connect(self.selectLabel)
        layout.addWidget(self.reference_button)

        # вывод выбранного эталона
        self.reference_image_label = QLabel()
        layout.addWidget(self.reference_image_label)

        """
        self.change_position_button = QPushButton("Изменить положение рамки вокруг эталонной картинки")
        self.change_position_button.clicked.connect(self.changePosition)
        self.change_position_button.setEnabled(False)  # Кнопка недоступна до выбора эталона
        layout.addWidget(self.change_position_button)
        """

        self.compare_button = QPushButton("Сравнить все изображения с эталоном")
        self.compare_button.setStyleSheet("background-color: #f0f0f0;"
                                          "border: 1px solid #ccc;"
                                          "color: black;"
                                          "border-radius: 10px;"
                                          "padding: 5px 10px;"
                                          "QPushButton:hover {"
                                          "background-color: #e0e0e0;"  # Цвет кнопки при наведении курсора
                                          "}"
                                          )
        self.compare_button.setFont(font)
        self.compare_button.clicked.connect(self.compareImages)
        self.compare_button.setEnabled(False)  # Начально блокируем кнопку
        layout.addWidget(self.compare_button)

        self.images_layout = QHBoxLayout()
        layout.addLayout(self.images_layout)



        # Виджет прокрутки
        self.scroll_area = QScrollArea()
        self.image_container = QWidget()
        self.image_layout = QVBoxLayout(self.image_container)
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setWidget(self.image_container)
        # layout.addWidget(self.scroll_area)


        self.setLayout(layout)
        self.setGeometry(500, 200, 700, 400)
        self.setWindowTitle('Image Comparison App')

    def openExcelFile(self):
        # Открыть Excel-файл
        QDesktopServices.openUrl(QUrl.fromLocalFile(self.excel_file))
    def selectFolder(self):
        """
        Обработчик нажатия кнопки для выбора папки
        """
        folder = QFileDialog.getExistingDirectory(self, "Выберите папку с изображениями")
        if folder:
            self.folder_path = folder
            self.folder_label.setText(f"Выбрана папка: {os.path.basename(folder)}")
            self.images_loaded = False
            self.enableCompareButtonIfReady()

    def selectLabel(self):
        """
        Обработчик нажатия кнопки для выбора и вывода эталона (вырезанная микросхема)
        """
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        label_path, _ = QFileDialog.getOpenFileName(self, "Выберите эталон", "", "Images (*.png *.jpg *.bmp *.jpeg);;All Files (*)", options=options)
        if label_path:
            self.label_path = label_path
            self.enableCompareButtonIfReady()
            self.reference_label.setText(f"Выбран эталон: {os.path.basename(os.path.splitext(label_path)[0])}")

            # Выводим изображение эталона
            etalon_img = cv2.imdecode(np.fromfile(label_path, dtype=np.uint8), cv2.IMREAD_GRAYSCALE)
            etalon_cut_data = get_window(etalon_img)['img_cutted']

            pixmap = self.convertCvImageToPixmap(etalon_cut_data, 300,300)
            self.reference_image_label.setPixmap(pixmap)
            self.images_loaded = False

    def enableCompareButtonIfReady(self):
        # Проверяем, выбраны ли папка и эталон, и разблокируем кнопку
        if self.folder_path and self.label_path:
            self.compare_button.setEnabled(True)
        else:
            self.compare_button.setEnabled(False)
    def convertCvImageToPixmap(self, cv_image, target_width=None, target_height=None):
        """
        Функция для преобразования cv-изображения в pixmap для вывода в окне приложения
        Задаем размер выходящего изображения (target_width, target_height)
        """

        if cv_image is not None:
            # Преобразование в RGB
            image_rgb = cv2.cvtColor(cv_image, cv2.COLOR_BGR2RGB)

            # Если указаны желаемые размеры вывода, изменяем размер изображения
            if target_width is not None and target_height is not None:
                image_rgb = cv2.resize(image_rgb, (target_width, target_height))

            height, width, channel = image_rgb.shape
            bytes_per_line = 3 * width

            # Создание объекта QImage
            q_image = QImage(image_rgb.data, width, height, bytes_per_line, QImage.Format_RGB888)

            '''
            # Применение colormap пока не работает!!!!
            colormap_image = q_image.convertToFormat(QImage.Format_Grayscale8)
            colormap_image.setColorTable([qRgb(i, i, i) for i in range(256)])
            '''


            # Преобразование QImage в QPixmap
            pixmap = QPixmap.fromImage(q_image)

            return pixmap
        else:
            print("Ошибка: Изображение не задано.")
            return None

    def update_excel(self, file_path, label_image_name, test_image_name, stat):
        """
        Функция для создания или обновления Excel-файла с указанными данными
        """
        folder_name = os.path.basename(self.folder_path)
        label_image_name = os.path.basename(self.label_path)
        print(folder_name, label_image_name, test_image_name, stat)

        try:
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(["Партия","Эталон", "Снимок", "Статус"])  # заголовки

            # Проверяем, есть ли уже такие данные в файле
            found = False
            for row in ws.iter_rows(min_row=2, max_col=4, max_row=ws.max_row):
                if row[0].value == folder_name and row[2].value == test_image_name:
                    row[3].value = stat
                    found = True
                    break

            if not found:
                ws.append([folder_name, label_image_name, test_image_name, stat])

            wb.save(file_path)
            return True
        except PermissionError:
            print("Ошибка: Файл Excel открыт. Пожалуйста, закройте файл и повторите попытку.")
            return False

    def clearImages(self):
        """
        Функция для очистки виджета отображения изображений
        """
        # Удаляем элементы из макета
        for i in reversed(range(self.images_layout.count())):
            widget = self.images_layout.itemAt(i).widget()
            if widget is not None:
                widget.setParent(None)

        # Сбрасываем флаг загрузки изображений
        self.images_loaded = False
    def compareImages(self):
        """
        Функция для сравнения изображений и вывода карты различий
        """
        # Проверяем, были ли изображения уже загружены
        if self.images_loaded:
            return
        else:
            self.clearImages()

        # Эталонное изображение
        label_image = cv2.imdecode(np.fromfile(self.label_path, dtype=np.uint8), cv2.IMREAD_GRAYSCALE)
        label_data = get_window(label_image)['img_cutted']

        # Фильтруем список файлов, исключая эталонное изображение
        files = os.listdir(self.folder_path)
        self.test_image_paths = [os.path.join(self.folder_path, file) for file in files if file != os.path.basename(self.label_path)]

        test_image = cv2.imdecode(np.fromfile(self.test_image_paths[0], dtype=np.uint8), cv2.IMREAD_GRAYSCALE)
        test_data = get_window(test_image)['img_cutted']


        superimpose_data = update_image_by_etalon(label_data,test_data)
        differences1_superimpose_data = find_differences1(label_data, superimpose_data['img_cutted'])


        # Преобразуем изображения cv в QPixmap
        test = self.convertCvImageToPixmap(test_data, 200, 200)
        ABS_with_SIFT = self.convertCvImageToPixmap(differences1_superimpose_data['img_diff'], 200, 200)
        # Erode_with_SIFT = self.convertCvImageToPixmap(differences1_superimpose_data['img_diff_erode'], 200, 200)
        # Dilate_with_SIFT = self.convertCvImageToPixmap(differences1_superimpose_data['img_diff_dilate'], 200, 200)

        titles = ["Test Image:", "ABS with SIFT:"]

        # Создаем три виджета QLabel для отображения изображений
        for image_cv, title in zip([test, ABS_with_SIFT], titles):
            # Создаем QLabel для заголовка
            title_label = QLabel(title)
            self.images_layout.addWidget(title_label)

            # Создаем QLabel для отображения изображения
            label = QLabel()
            label.setPixmap(image_cv)
            self.images_layout.addWidget(label)

        self.update_excel(os.path.normpath(self.excel_file), os.path.basename(self.label_path),os.path.basename(self.test_image_paths[0]), "Различий нет")
        # Создаем чекбокс
        checkbox = QCheckBox("Изображения различаются?")
        checkbox.stateChanged.connect(lambda state, label_image=os.path.basename(self.label_path), test_image=os.path.basename(self.test_image_paths[0]): self.update_excel(os.path.normpath(self.excel_file), label_image, test_image, "Найдены различия!") if state == Qt.Checked else self.update_excel(os.path.normpath(self.excel_file), label_image, test_image, "Различий нет"))


        self.images_layout.addWidget(checkbox)


        self.images_layout.parentWidget().show()
        self.images_loaded = True  # изображения загружены



    def changePosition(self):
        # Изменение положения рамки вокруг эталонной картинки
        # Этот код будет вызываться при нажатии кнопки "Изменить положение рамки"

        # Пока не реализовано
        self.result_group_box.show()


    def displayComparisonResults(self):
        # Очищаем предыдущие результаты
        for i in reversed(range(self.image_layout.count())):
            widget = self.image_layout.takeAt(i).widget()
            if widget is not None:
                widget.setParent(None)

        # Выводим результаты
        for result_image in self.result_images:
            result_pixmap = self.imageToPixmap(result_image)
            result_label = QLabel()
            result_label.setPixmap(result_pixmap)
            result_label.setAlignment(Qt.AlignCenter)
            self.image_layout.addWidget(result_label)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = ImageComparisonApp()
    ex.show()
    sys.exit(app.exec_())
